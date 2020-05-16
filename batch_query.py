#coding=utf-8
import requests
import ConfigParser
import re
import os
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from HTMLParser import HTMLParser
#import json


def remove_BOM(config_path):
	content = open(config_path).read()
	content = re.sub(r"\xfe\xff","", content)
	content = re.sub(r"\xff\xfe","", content)
	content = re.sub(r"\xef\xbb\xbf","", content)
	open(config_path, 'w').write(content)

remove_BOM("config.ini")
cf = ConfigParser.ConfigParser()
cf.read("config.ini")
#excelFile = cf.get("config", "excelFile")
inputFieldName = cf.get("config", "inputFieldName")
exportFieldNames = cf.get("config", "exportFieldNames").replace(' ', '').split(',')
proxyTest = True if cf.get("config", "proxyTest") == "1" else False
proxyHeader = cf.get("config", "proxyHeader").replace(' ', '').split(',')
proxy = cf.get("config", "proxy")

#print ("excelFile:" + excelFile).decode('utf-8')
print ("proxyTest:" + ("1" if proxyTest else "0"))
print ("inputFieldName:" + inputFieldName).decode('utf-8')
print ("exportFieldNames:" + cf.get("config", "exportFieldNames")).decode('utf-8')

proxiesList = []
if proxyTest:
	with open("proxy.txt", "r") as f:
		for line in f.readlines():
			line = line.strip('\n').replace('PROXY ', '').replace('"', '')
			line = re.sub('\s+', '', line).strip()
			proxiesList.append(line)
else:
	proxiesList.append(proxy);

get_now_milli_time = lambda: int(time.time() * 1000)

recordStartTime = 0
proxyTimeList = {}
for p in range(len(proxiesList)):
	if proxyTest == True:
		recordStartTime = get_now_milli_time()
	proxies = {}
	proxy = proxiesList[p]
	if not proxy == "":
		for i in range(len(proxyHeader)):
			proxies[proxyHeader[i]] = proxyHeader[i] + "://" + proxy
		#proxies={
		#	'http':'http://'+proxy,
		#	'https':'https://'+proxy,
		#}
		
	print proxies
	headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
	'Accept-Encoding': 'gzip, deflate',
	'Accept-Language': 'zh-CN,zh;q=0.9',
	'Connection': 'keep-alive',
	'Host': 'www.yesinfo.com.cn',
	'Upgrade-Insecure-Requests': '1',
	'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.129 Safari/537.36'}
	try:
		session = requests.Session()
		html = session.get('http://www.yesinfo.com.cn/', headers=headers, proxies=proxies)
	except:
		if proxyTest == True:
			proxyTimeList[proxy] = str(get_now_milli_time() - recordStartTime) + "#Fail"
		continue;

	headers["Origin"]="http://www.yesinfo.com.cn"
	headers["Referer"]="http://www.yesinfo.com.cn/"


	pattern = re.compile('[^~].*?\.(xlsx)$')
	files = os.listdir('./')
	for file in files:
		if pattern.match(file):
			wb = load_workbook(file, data_only=True)
			ws = wb.worksheets[0]

			exportIndex = 1;
			totalNum = 0
			if not inputFieldName == "":
				for i in range(1,ws.max_row + 1):
					if ws.cell(row=1, column=i).value.lower().find(inputFieldName.lower()) != -1:
						exportIndex = i
						for k in range(2, ws.max_row+1):
							cntrId = ws.cell(row=k, column=exportIndex).value
							if cntrId == None or cntrId.strip() == "":
								continue
							totalNum = totalNum + 1
						break
			else:
				for k in range(2, ws.max_row+1):
					cntrId = ws.cell(row=k, column=1).value
					if cntrId == None or cntrId.strip() == "":
						continue
					totalNum = totalNum + 1

			print ('###########################################')
			print ('Excel名称:'+ str(file)).decode('utf-8')
			print ('导出列号:'+ str(exportIndex)).decode('utf-8')
			print ('查询数量:'+ str(totalNum)).decode('utf-8')

			for i in range(0,len(exportFieldNames)):
				ws.cell(row=1, column=exportIndex + 1 + i).value = exportFieldNames[i].decode('utf-8')
			ws.cell(row=1, column=exportIndex + 1 +len(exportFieldNames)).value = "查询状态".decode('utf-8')


			print ('###########################################')
			print ('开始查询').decode('utf-8')

			curNum=1;
			for k in range(2, ws.max_row+1):
				cntrId = ws.cell(row=k, column=exportIndex).value
				if cntrId == None or cntrId.strip() == "":
					continue
				print ("正在查询"+str(curNum)+"/"+str(totalNum)).decode('utf-8')
				curNum = curNum + 1
				print cntrId
				try:
					html = session.post('http://www.yesinfo.com.cn/homepage/publicInquiry/contInquiry.action', data={'cntrId':cntrId}, headers=headers, proxies=proxies)
					#text = html.text.encode('gbk', 'ignore').decode('gbk')
					text = html.text.split('jzDetails jzDetails_js">')[1]
					text = text.replace('<ul>', '')
					text = text.split('</ul>')[0]
					text = text.replace(' ', '')
					text = text.replace('\r','').replace('\n','').replace('\t','')
					textArray = text.split('</li>')
					vlen = len(textArray)
					dataKV = {}
					for i in range(0,vlen):
						temp = textArray[i].split('</div>')
						if len(temp) >= 2:
							dataKV[temp[0].replace('<li><divclass="fl">', '').replace('：'.decode('utf-8'), '')] = temp[1].replace('<divclass="fr">', "")
					for i in range(0,len(exportFieldNames)):
						print exportFieldNames[i].decode('utf-8')
						print HTMLParser().unescape(dataKV[exportFieldNames[i].decode('utf-8')])
						ws.cell(row=k, column=exportIndex + 1+i).value = HTMLParser().unescape(dataKV[exportFieldNames[i].decode('utf-8')])
					ws.cell(row=k, column=exportIndex + 1+len(exportFieldNames)).value = "成功".decode('utf-8')
				except:
					for i in range(0,len(exportFieldNames)):
						print exportFieldNames[i].decode('utf-8')
						print ''
						ws.cell(row=k, column=exportIndex + 1+i).value = ''
					ws.cell(row=k, column=exportIndex + 1+len(exportFieldNames)).value = "失败".decode('utf-8')
			saved = False
			while saved == False:
				try:		
					wb.save(file)
					saved = True
					print '保存excel成功！'.decode('utf-8')
				except:
					print '保存excel失败，请关闭excel后按任意键重试！'.decode('utf-8')
					raw_input()
			if proxyTest == True:
				proxyTimeList[proxy] = str(get_now_milli_time() - recordStartTime) + "#Success"

if proxyTest == True:
	proxyRet = ''
	for key in proxyTimeList:
		proxyRet = proxyRet + key + "," + proxyTimeList[key] + '\n'
	with open("proxy_test_result.txt", "a") as f:
		f.write(proxyRet)
print "#############按任意键退出#############".decode('utf-8')
raw_input()